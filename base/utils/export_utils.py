# coding: utf8
##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2016 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Color, Style, PatternFill
from django.utils.translation import ugettext_lazy as _

from base import models as mdl


HEADER = [str(_('Academic year')),
          str(_('Session')),
          str(_('Activity code')),
          str(_('Program')),
          str(_('Registration number')),
          str(_('Last name')),
          str(_('First name')),
          str(_('Numbered score')),
          str(_('Other score')),
          str(_('End date')),
          str(_('ID'))]


def export_xls(request, session_id, academic_year_id):

    academic_year = mdl.academic_year.find_academic_year_by_id(academic_year_id)
    session_exam = mdl.session_exam.find_session_by_id(session_id)
    is_fac = mdl.program_manager.is_programme_manager(request.user,session_exam.offer_year_calendar.offer_year)
    wb = Workbook()
    ws = wb.active

    __columns_ajusting(ws)

    # masquage de la colonne avec l'id exam enrollment

    ws.append(HEADER)

    dv = __create_data_list_for_justification(is_fac)
    ws.add_data_validation(dv)

    cptr = 1
    for rec_exam_enrollment in mdl.exam_enrollment.find_exam_enrollments_by_session(session_exam):
        student = rec_exam_enrollment.learning_unit_enrollment.student
        o = rec_exam_enrollment.learning_unit_enrollment.offer
        person = mdl.person.find_person(student.person.id)

        if session_exam.offer_year_calendar.end_date is None:
            end_date = "-"
        else:
            end_date = session_exam.offer_year_calendar.end_date.strftime('%d/%m/%Y')
        score = None
        if rec_exam_enrollment.score_final:
            if rec_exam_enrollment.session_exam.learning_unit_year.decimal_scores:
                score = "{0:.2f}".format(rec_exam_enrollment.score_final)
            else:
                score = "{0:.0f}".format(rec_exam_enrollment.score_final)
        justification = ""
        if rec_exam_enrollment.justification_final:
            justification = dict(mdl.exam_enrollment.JUSTIFICATION_TYPES)[rec_exam_enrollment.justification_final]
        ws.append([str(academic_year),
                   str(session_exam.number_session),
                   session_exam.learning_unit_year.acronym,
                   o.acronym,
                   student.registration_id,
                   person.last_name,
                   person.first_name,
                   score,
                   str(justification),
                   end_date,
                   rec_exam_enrollment.id])

        cptr += 1
        __coloring_non_editable(ws, cptr, rec_exam_enrollment.encoding_status,score,rec_exam_enrollment.justification_final)

    # Ajouter 100 pour si on ajoute des enregistrements
    dv.ranges.append('I2:I' + str(cptr + 100))

    response = HttpResponse(content=save_virtual_workbook(wb))
    response['Content-Disposition'] = 'attachment; filename=score_encoding.xlsx'
    response['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


def __columns_ajusting(ws):
    """
    ajustement des colonnes à la bonne dimension
    """
    col_academic_year = ws.column_dimensions['A']
    col_academic_year.width = 18
    col_academic_year = ws.column_dimensions['C']
    col_academic_year.width = 18
    col_academic_year = ws.column_dimensions['E']
    col_academic_year.width = 18
    col_last_name = ws.column_dimensions['F']
    col_last_name.width = 30
    col_first_name = ws.column_dimensions['G']
    col_first_name.width = 30
    col_note = ws.column_dimensions['H']
    col_note.width = 20
    col_note = ws.column_dimensions['I']
    col_note.width = 20
    col_id_exam_enrollment = ws.column_dimensions['K']
    col_id_exam_enrollment.hidden = True


def __create_data_list_for_justification(is_fac):
    """
    Création de la liste de choix pour la justification
    """
    dv = DataValidation(type="list", formula1='"%s"' % mdl.exam_enrollment.justification_label_authorized(is_fac),
                        allow_blank=True)
    dv.error = str(_('Invalid entry, not in the list of choices'))
    dv.errorTitle = str(_('Invalid entry'))

    dv.prompt = str(_('Please choose in the list'))
    dv.promptTitle = str(_('List of choices'))
    return dv


def __coloring_non_editable(ws, cptr, encoding_status, score, justification):
    """
    définition du style pour les colonnes qu'on ne doit pas modifier
    """
    style_no_modification = Style(fill=PatternFill(patternType='solid', fgColor=Color('C1C1C1')))

    # coloration des colonnes qu'on ne doit pas modifier
    i = 1
    while i < 12:
        if i < 8 or i > 9:
            ws.cell(row=cptr, column=i).style = style_no_modification
        else:
            if not(score is None and justification is None):
                ws.cell(row=cptr, column=8).style = style_no_modification
                ws.cell(row=cptr, column=9).style = style_no_modification

        i += 1
